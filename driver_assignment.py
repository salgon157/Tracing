"""
driver_assignment.py — celodenní přiřazení řidičů k naplánovaným linkám
=======================================================================

SAMOSTATNÝ krok PO naplánování všech dep dne (do plan_day se nezapojuje;
spouštění vyřeší vrstva nad námi):

  python driver_assignment.py 2026-08-13
  python driver_assignment.py 2026-08-13 --label b5      # testovací běhy
  python driver_assignment.py 2026-08-13 --depots CB MO  # vědomá podmnožina

Vstupy:
  data/ridici/aktivni/*.xlsx            registr aut+řidičů z ESO (právě jeden;
                                        PII — složka je gitignored)
  data/results/{DEPO}/{DATUM}/          lines_summary.csv + lines_stops.csv
                                        všech dep dne

Výstupy:
  data/results/driver_assignment/{DATUM}/driver_plan_{DATUM}.csv  (celý den)
  data/results/{DEPO}/{DATUM}/driver_plan_{DEPO}_{DATUM}.csv      (per depo)
  data/results/driver_assignment/{DATUM}/summary.json

Model: JEDNA přiřazovací úloha za celý den — všechny linky všech dep ×
řidiči. Globální optimum najde maďarský algoritmus; CB si tak „nevyžere"
řidiče, kteří se víc hodí na HK linky.

HARD (zakázaná buňka): den v týdnu nesedí / Dostupnost≠Ano / Aktivní≠Ano /
řidič nemá auto správného typu. Řidič jede max jednu linku denně (i když
má víc aut). Dvojlinka (2 linky téhož vozidla) = jedna jednotka.

SOFT (skóre 0–1 × váha z CONFIG, viz jednotlivé funkce):
  plneni_planu — kdo zaostává za poměrnou částí ročního plánu km
                 (BEZ DAT dokud ESO neplní Aktual. km — pak neutrální)
  dojezd       — dlouhé linky vzdáleným řidičům (pořadové párování)
  kvalita_tightness — Rychlý na linky s napjatými okny; tightness na
                 konci linky váží víc než na začátku
  familiarity  — podíl zastávek, které řidič zná z historie závozů
                 (BEZ DAT dokud historie nenese sloupec řidiče)
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

from fleet_budget import DEPOT_ORDER

# ── Konfigurace (váhy a parametry — ladí se tady, ne v kódu) ─────────────────
CONFIG = {
    "weights": {
        "plneni_planu":      3.0,   # priorita: smluvní km se musí plnit
        "dojezd":            1.0,
        "kvalita_tightness": 1.0,
        "familiarity":       1.0,   # zatím bez dat (historie nemá řidiče)
    },
    # Zastávka je "tight", když rezerva do KONCE původního okna je <= tolik
    # minut (příjezd po konci okna — v toleranci +25 — je tight vždy).
    "tight_slack_min": 15,
    # Pozice na lince: zastávka na konci váží (1 + koef)x víc než na začátku.
    # 0.3 drží vlastnost: 5 tight na konci (6.5) > 5 na začátku (5),
    # ale < 7 na začátku (7).
    "tight_pos_coef": 0.3,
    # Kvalita řidiče -> "rychlost" 0..1; skóre = 1 - |tightness - rychlost|
    "quality_speed": {"Rychlý": 1.0, "Standart": 0.5, "Pomalý": 0.0},
    "ridici_dir":   "data/ridici/aktivni",
    "results_root": "data/results",
    "history_dir":  "data/historie_objednavky",
}

BIG_COST = 1e9          # zakázaná buňka (hard constraint)

# (Typ vozidla, Nosnost) z ESO registru -> náš TYPE kód (dle vehicle_types)
TYPE_BY_TYP_NOSNOST = {
    ("do 3t", 1200):  "TYPE_01",
    ("do 3t", 1350):  "TYPE_02",
    ("do 7t", 3000):  "TYPE_03",
    ("do 7t", 3200):  "TYPE_04",
    ("do 18t", 8000): "TYPE_05",
    ("do 18t", 8700): "TYPE_06",
    ("do 4t", 2000):  "TYPE_07",
}

DAY_NAMES = ["Po", "Út", "St", "Čt", "Pá", "So", "Ne"]   # weekday() 0..6


# ═════════════════════════════════════════════════════════════════════════════
#  Registr aut a řidičů (xlsx z ESO)
# ═════════════════════════════════════════════════════════════════════════════

def parse_days(spec: str) -> set[int]:
    """
    'Dny použitelnosti vozidla' -> množina dní (Po=0 … Ne=6).

    Lomítko dělí týdenní a víkendovou část, obě se sjednocují:
      'Po-Pá/So-Ne'      -> {0..6}
      'Po,Pá'            -> {0, 4}
      'Po,St,Pá/So-Ne'   -> {0, 2, 4, 5, 6}
      'Po/So-Ne'         -> {0, 5, 6}
    Neznámý zápis = chyba (tichá dezinterpretace by řidiče nasadila
    v den, kdy nejezdí).
    """
    days: set[int] = set()
    for part in str(spec).split("/"):
        for token in part.split(","):
            token = token.strip()
            if not token:
                continue
            if "-" in token:
                a, b = (t.strip() for t in token.split("-", 1))
                if a not in DAY_NAMES or b not in DAY_NAMES:
                    raise ValueError(f"[CHYBA] Neznámý rozsah dní {token!r} v {spec!r}")
                days.update(range(DAY_NAMES.index(a), DAY_NAMES.index(b) + 1))
            else:
                if token not in DAY_NAMES:
                    raise ValueError(f"[CHYBA] Neznámý den {token!r} v {spec!r}")
                days.add(DAY_NAMES.index(token))
    if not days:
        raise ValueError(f"[CHYBA] Prázdné dny použitelnosti: {spec!r}")
    return days


def map_type(typ: str, nosnost) -> str:
    key = (str(typ).strip(), int(float(nosnost or 0)))
    if key not in TYPE_BY_TYP_NOSNOST:
        raise ValueError(
            f"[CHYBA] Neznámá kombinace typu a nosnosti v registru: {key}. "
            f"Známé: {sorted(TYPE_BY_TYP_NOSNOST)}")
    return TYPE_BY_TYP_NOSNOST[key]


def find_registry_file(ridici_dir: Path | str | None = None) -> Path:
    """Právě jeden xlsx v data/ridici/aktivni — víc/míň je chyba (stejná
    filozofie jako riro aktivni/: program mezi soubory nevybírá)."""
    d = Path(ridici_dir if ridici_dir is not None else CONFIG["ridici_dir"])
    files = sorted(d.glob("*.xlsx")) if d.exists() else []
    if len(files) != 1:
        raise SystemExit(
            f"[CHYBA] V {d.as_posix()}/ musí být právě jeden .xlsx s registrem "
            f"aut a řidičů z ESO (nalezeno {len(files)}).\n"
            f"        Nahraj export 'Auta - Řidiči - Eso.xlsx' do téhle složky.")
    return files[0]


def _num(v) -> float | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_registry(path: Path) -> list[dict]:
    """Použitelné řádky registru (Použít vozidlo=Ano), per VOZIDLO."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else "" for h in next(rows_iter)]

    required = ["Číslo vozidla", "Jméno řidiče", "Typ vozidla [23]",
                "Nosnost vozidla [5]", "Dny použitelnosti vozidla",
                "Dostupnost vozidla [8]", "Aktivní", "Použít vozidlo",
                "Km dojezd depo", "Kvalita řidiče"]
    missing = [c for c in required if c not in header]
    if missing:
        raise SystemExit(f"[CHYBA] Registru chybí sloupce: {missing} — "
                         f"jiný formát exportu z ESO?")

    idx = {h: i for i, h in enumerate(header)}

    def col(row, name, default=None):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else default

    out = []
    for row in rows_iter:
        if not row or not col(row, "Číslo vozidla"):
            continue
        if str(col(row, "Použít vozidlo", "")).strip() != "Ano":
            continue
        out.append({
            "vehicle_no":   str(col(row, "Číslo vozidla")).strip(),
            "vehicle_name": str(col(row, "název vozidla") or "").strip(),
            "driver":       str(col(row, "Jméno řidiče") or "").strip(),
            "dopravce":     str(col(row, "Název dopravce") or "").strip(),
            "type_code":    map_type(col(row, "Typ vozidla [23]"),
                                     col(row, "Nosnost vozidla [5]")),
            "days":         parse_days(col(row, "Dny použitelnosti vozidla")),
            "available":    str(col(row, "Dostupnost vozidla [8]", "")).strip() == "Ano",
            "active":       str(col(row, "Aktivní", "")).strip() == "Ano",
            "dojezd_km":    _num(col(row, "Km dojezd depo")) or 0.0,
            "kvalita":      str(col(row, "Kvalita řidiče") or "Standart").strip(),
            "plan_rok":     _num(col(row, "Plán km rok")),
            "plan_mes":     _num(col(row, "Plán km měs.")),
            "aktual_rok":   _num(col(row, "Aktual. km rok")),
            "aktual_mes":   _num(col(row, "Aktual. km měs.")),
        })
    if not out:
        raise SystemExit(f"[CHYBA] Registr {path} nemá žádné použitelné řádky "
                         f"(Použít vozidlo=Ano).")
    return out


# ═════════════════════════════════════════════════════════════════════════════
#  Linky dne (výsledky solveru všech dep)
# ═════════════════════════════════════════════════════════════════════════════

def _time_to_min(hhmm: str) -> int | None:
    try:
        h, m = hhmm.strip().split(":")
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return None


def load_depot_lines(results_dir: Path, depot: str) -> list[dict]:
    """Linky depa vč. zastávek. Dvojlinky (společné vehicle_id) sloučené
    do jedné jednotky — jeden řidič jede obě jízdy."""
    summary = results_dir / "lines_summary.csv"
    stops_f = results_dir / "lines_stops.csv"
    for f in (summary, stops_f):
        if not f.exists():
            raise SystemExit(f"[CHYBA] Chybí {f} — depo {depot} nemá "
                             f"kompletní výsledky.")

    stops_by_line: dict[str, list[dict]] = {}
    with open(stops_f, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if not str(row.get("order_id", "")).strip():
                continue        # sklad (start/návrat)
            win = str(row.get("window", ""))
            end = _time_to_min(win.split("–")[-1]) if "–" in win else None
            arr = _time_to_min(str(row.get("arrival", "")))
            stops_by_line.setdefault(row["line_id"], []).append({
                "location_code": str(row.get("location_code", "")).strip(),
                "arrival_min":   arr,
                "window_end_min": end,
            })

    by_vehicle: dict[str, dict] = {}
    with open(summary, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            vehicle_id = str(row.get("vehicle_id", "")).strip()
            if not vehicle_id:
                continue        # CELKEM
            line_id = str(row.get("line_id", "")).strip()
            unit = by_vehicle.setdefault(vehicle_id, {
                "depot": depot,
                "vehicle_id": vehicle_id,
                "type_code": vehicle_id.rsplit("_", 1)[0],
                "line_ids": [], "km": 0.0, "tightness_raw": 0.0,
                "stops_total": 0, "locations": set(),
            })
            unit["line_ids"].append(line_id)
            unit["km"] += float(row.get("total_km", 0) or 0)
            stops = stops_by_line.get(line_id, [])
            unit["tightness_raw"] += line_tightness(stops)
            unit["stops_total"] += len(stops)
            unit["locations"].update(s["location_code"] for s in stops)
    return list(by_vehicle.values())


def line_tightness(stops: list[dict],
                   slack_min: int | None = None,
                   pos_coef: float | None = None) -> float:
    """
    T = Σ tightᵢ × (1 + pos_coef × posᵢ),  posᵢ ∈ [0, 1] pozice na lince.

    Tight = rezerva do konce PŮVODNÍHO okna <= slack_min (příjezd po konci
    okna je tight vždy). Konec linky váží víc než začátek (zpoždění se
    kumuluje), ale pozice nesmí přebít počet — proto koeficient, ne násobek.
    """
    slack_min = CONFIG["tight_slack_min"] if slack_min is None else slack_min
    pos_coef = CONFIG["tight_pos_coef"] if pos_coef is None else pos_coef
    n = len(stops)
    total = 0.0
    for i, s in enumerate(stops):
        if s["arrival_min"] is None or s["window_end_min"] is None:
            continue
        if s["window_end_min"] - s["arrival_min"] <= slack_min:
            pos = i / (n - 1) if n > 1 else 1.0
            total += 1.0 + pos_coef * pos
    return total


# ═════════════════════════════════════════════════════════════════════════════
#  Skóre (každé 0..1, vyšší = lepší pár)
# ═════════════════════════════════════════════════════════════════════════════

def percentile_ranks(values: list[float]) -> list[float]:
    """Pořadová normalizace do [0,1]; shodné hodnoty = shodný rank
    (deterministicky). Jedna hodnota -> 0.5."""
    n = len(values)
    if n <= 1:
        return [0.5] * n
    return [(sum(1 for v in values if v < x)
             + (sum(1 for v in values if v == x) - 1) / 2) / (n - 1)
            for x in values]


def plan_deficit(row: dict, day_of_year: int) -> float | None:
    """Relativní skluz vůči poměrné části ročního plánu; None = bez dat.
    (Až ESO začne plnit Aktual./Plán měs., přimíchá se tady.)"""
    if not row["plan_rok"] or row["aktual_rok"] is None:
        return None
    expected = row["plan_rok"] * day_of_year / 365.0
    return (expected - row["aktual_rok"]) / row["plan_rok"]


def load_familiarity(history_dir: Path | str | None = None) -> dict | None:
    """{location_code -> set(řidičů)} z historie závozů — AŽ historie
    ponese sloupec řidiče (dnes nenese -> None = kritérium bez dat)."""
    d = Path(history_dir if history_dir is not None else CONFIG["history_dir"])
    files = sorted(d.glob("*.xlsx")) if d.exists() else []
    if not files:
        return None
    import openpyxl
    ws = openpyxl.load_workbook(files[-1], read_only=True, data_only=True).active
    header = [str(h).strip() if h else "" for h in next(ws.iter_rows(values_only=True))]
    driver_cols = [h for h in header if "idi" in h.lower()]   # Řidič/ridic…
    if not driver_cols:
        return None
    zk, dr = header.index("Zkratka"), header.index(driver_cols[0])
    fam: dict[str, set] = {}
    for f in files:
        ws = openpyxl.load_workbook(f, read_only=True, data_only=True).active
        rows = ws.iter_rows(values_only=True)
        next(rows)
        for r in rows:
            loc = str(r[zk] or "").strip().lower()
            drv = str(r[dr] or "").strip()
            if loc and drv:
                fam.setdefault(loc, set()).add(drv)
    return fam or None


# ═════════════════════════════════════════════════════════════════════════════
#  Sestavení matice a přiřazení
# ═════════════════════════════════════════════════════════════════════════════

def build_assignment(units: list[dict], registry: list[dict], target_date: str,
                     familiarity: dict | None = None,
                     weights: dict | None = None) -> dict:
    """
    Celodenní matice jednotky × ŘIDIČI; buňka = nejlepší řidičovo auto
    správného typu. Vrací {"assigned": [...], "uncovered": [...],
    "warnings": [...]}.
    """
    W = dict(CONFIG["weights"]) if weights is None else dict(weights)
    dt = datetime.strptime(target_date, "%Y-%m-%d")
    weekday, doy = dt.weekday(), dt.timetuple().tm_yday
    warnings: list[str] = []

    units = sorted(units, key=lambda u: (u["depot"], u["vehicle_id"]))
    drivers = sorted({r["driver"] for r in registry})
    rows_by_driver: dict[str, list[dict]] = {}
    for r in registry:
        rows_by_driver.setdefault(r["driver"], []).append(r)

    # Pořadové normalizace přes CELÝ den / celý registr
    km_rank = percentile_ranks([u["km"] for u in units])
    t_rank = percentile_ranks([u["tightness_raw"] for u in units])
    all_rows = [r for rs in rows_by_driver.values() for r in rs]
    dojezd_rank = dict(zip(map(id, all_rows),
                           percentile_ranks([r["dojezd_km"] for r in all_rows])))
    deficits = [plan_deficit(r, doy) for r in all_rows]
    with_data = [d for d in deficits if d is not None]
    if not with_data:
        warnings.append("[!] Plnění plánu BEZ DAT — ESO zatím neplní "
                        "Aktual. km; kritérium je neutrální (0.5).")
        deficit_rank = {id(r): 0.5 for r in all_rows}
    else:
        ranks = percentile_ranks(with_data)
        it = iter(ranks)
        deficit_rank = {id(r): (next(it) if d is not None else 0.5)
                        for r, d in zip(all_rows, deficits)}
    if familiarity is None:
        warnings.append("[!] Familiarity BEZ DAT — historie závozů nenese "
                        "řidiče; kritérium je neutrální (0.5).")
    speed = CONFIG["quality_speed"]

    def cell(u_idx: int, driver: str) -> tuple[float, dict] | None:
        """Nejlepší (skóre, auto) řidiče pro jednotku; None = hard zákaz."""
        u = units[u_idx]
        best = None
        for r in rows_by_driver[driver]:
            if (r["type_code"] != u["type_code"] or weekday not in r["days"]
                    or not r["available"] or not r["active"]):
                continue
            s_plan = deficit_rank[id(r)]
            s_doj = 1.0 - abs(km_rank[u_idx] - dojezd_rank[id(r)])
            s_qual = 1.0 - abs(t_rank[u_idx]
                               - speed.get(r["kvalita"], 0.5))
            if familiarity is None or not u["locations"]:
                s_fam = 0.5
            else:
                known = sum(1 for loc in u["locations"]
                            if driver in familiarity.get(loc, ()))
                s_fam = known / len(u["locations"])
            score = (W["plneni_planu"] * s_plan + W["dojezd"] * s_doj
                     + W["kvalita_tightness"] * s_qual
                     + W["familiarity"] * s_fam)
            entry = (score, {"row": r,
                             "breakdown": {"plneni": round(s_plan, 3),
                                           "dojezd": round(s_doj, 3),
                                           "kvalita": round(s_qual, 3),
                                           "familiarity": round(s_fam, 3)}})
            if best is None or entry[0] > best[0]:
                best = entry
        return best

    cells: dict[tuple[int, int], tuple[float, dict]] = {}
    cost = [[BIG_COST] * len(drivers) for _ in units]
    for ui in range(len(units)):
        for di, drv in enumerate(drivers):
            c = cell(ui, drv)
            if c is not None:
                cells[(ui, di)] = c
                cost[ui][di] = -c[0]

    from scipy.optimize import linear_sum_assignment
    import numpy as np
    row_ind, col_ind = linear_sum_assignment(np.array(cost))

    assigned, uncovered = [], []
    matched = {ui: di for ui, di in zip(row_ind, col_ind)
               if (ui, di) in cells}
    for ui, u in enumerate(units):
        if ui not in matched:
            uncovered.append(u)
            continue
        di = matched[ui]
        score, info = cells[(ui, di)]
        assigned.append({"unit": u, "driver": drivers[di],
                         "vehicle": info["row"], "score": round(score, 3),
                         "breakdown": info["breakdown"]})
    if uncovered:
        warnings.append(f"[ALERT] {len(uncovered)} linek BEZ řidiče — po "
                        f"tvrdých filtrech (den, dostupnost, typ) nezbyl "
                        f"nikdo. Ruční zásah nutný.")
    return {"assigned": assigned, "uncovered": uncovered,
            "warnings": warnings, "weights": W}


# ═════════════════════════════════════════════════════════════════════════════
#  Výstupy
# ═════════════════════════════════════════════════════════════════════════════

CSV_HEADER = ["depot", "line_id", "vehicle_id", "type_code", "km",
              "tightness", "driver", "vehicle_no", "vehicle_name",
              "dopravce", "kvalita", "dojezd_km", "score",
              "s_plneni", "s_dojezd", "s_kvalita", "s_familiarity",
              "dvojlinka"]


def result_rows(result: dict) -> list[dict]:
    """Jeden řádek per LINKA (dvojlinka -> 2 řádky, týž řidič)."""
    rows = []
    for a in sorted(result["assigned"],
                    key=lambda x: (x["unit"]["depot"], x["unit"]["vehicle_id"])):
        u, v, b = a["unit"], a["vehicle"], a["breakdown"]
        for line_id in u["line_ids"]:
            rows.append({
                "depot": u["depot"], "line_id": line_id,
                "vehicle_id": u["vehicle_id"], "type_code": u["type_code"],
                "km": round(u["km"], 1),
                "tightness": round(u["tightness_raw"], 2),
                "driver": a["driver"], "vehicle_no": v["vehicle_no"],
                "vehicle_name": v["vehicle_name"], "dopravce": v["dopravce"],
                "kvalita": v["kvalita"], "dojezd_km": v["dojezd_km"],
                "score": a["score"], "s_plneni": b["plneni"],
                "s_dojezd": b["dojezd"], "s_kvalita": b["kvalita"],
                "s_familiarity": b["familiarity"],
                "dvojlinka": "ano" if len(u["line_ids"]) > 1 else "",
            })
    for u in result["uncovered"]:
        for line_id in u["line_ids"]:
            rows.append({"depot": u["depot"], "line_id": line_id,
                         "vehicle_id": u["vehicle_id"],
                         "type_code": u["type_code"], "km": round(u["km"], 1),
                         "tightness": round(u["tightness_raw"], 2),
                         "driver": "!!! NEPŘIŘAZENO !!!", "vehicle_no": "",
                         "vehicle_name": "", "dopravce": "", "kvalita": "",
                         "dojezd_km": "", "score": "", "s_plneni": "",
                         "s_dojezd": "", "s_kvalita": "", "s_familiarity": "",
                         "dvojlinka": "ano" if len(u["line_ids"]) > 1 else ""})
    return rows


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_HEADER)
        w.writeheader()
        w.writerows(rows)


def format_table(rows: list[dict]) -> str:
    out = []
    for depot in sorted({r["depot"] for r in rows}):
        out.append(f"\n{depot}:")
        out.append(f"  {'linka':<8} {'typ':<8} {'km':>7} {'tight':>6}  "
                   f"{'řidič':<24} {'auto':<6} {'kvalita':<9} {'skóre':>6}")
        for r in [x for x in rows if x["depot"] == depot]:
            out.append(f"  {r['line_id']:<8} {r['type_code']:<8} "
                       f"{r['km']:>7} {r['tightness']:>6}  "
                       f"{r['driver']:<24} {r['vehicle_no']:<6} "
                       f"{r['kvalita']:<9} {r['score']!s:>6}"
                       + ("  [dvojlinka]" if r["dvojlinka"] else ""))
    return "\n".join(out)


# ═════════════════════════════════════════════════════════════════════════════
#  Main
# ═════════════════════════════════════════════════════════════════════════════

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Celodenní přiřazení řidičů k naplánovaným linkám "
                    "(spouštět až po naplánování VŠECH dep dne).")
    ap.add_argument("date", help="Datum závozu (YYYY-MM-DD)")
    ap.add_argument("--depots", nargs="*", default=None,
                    help=f"Vědomá podmnožina dep (default: všechna "
                         f"{' '.join(DEPOT_ORDER)} — celodenní optimum "
                         f"potřebuje celý den)")
    ap.add_argument("--label", default="",
                    help="Přípona výsledkových složek (testovací běhy)")
    ap.add_argument("--results-root", default=CONFIG["results_root"])
    ap.add_argument("--ridici-file", default="",
                    help="Explicitní cesta k registru (default: jediný xlsx "
                         f"v {CONFIG['ridici_dir']}/)")
    args = ap.parse_args()

    depots = [d.upper() for d in args.depots] if args.depots else list(DEPOT_ORDER)
    unknown = sorted(set(depots) - set(DEPOT_ORDER))
    if unknown:
        raise SystemExit(f"[CHYBA] Neznámá depa: {unknown}")
    suffix = f"_{args.label}" if args.label else ""
    root = Path(args.results_root)

    registry_path = (Path(args.ridici_file) if args.ridici_file
                     else find_registry_file())
    registry = load_registry(registry_path)

    print("=" * 66)
    print(f"PŘIŘAZENÍ ŘIDIČŮ — {args.date} "
          f"({DAY_NAMES[datetime.strptime(args.date, '%Y-%m-%d').weekday()]})"
          + (f" | label {args.label}" if args.label else ""))
    print(f"Registr: {registry_path.name} — {len(registry)} aut, "
          f"{len({r['driver'] for r in registry})} řidičů")
    print("=" * 66)

    missing = [d for d in depots
               if not (root / d / f"{args.date}{suffix}" / "lines_summary.csv").exists()]
    if missing:
        raise SystemExit(
            f"[CHYBA] Chybí výsledky dep: {', '.join(missing)} "
            f"(hledám {root.as_posix()}/{{DEPO}}/{args.date}{suffix}/).\n"
            f"        Celodenní optimum potřebuje všechna depa — nejdřív "
            f"doplánuj, nebo vědomě: --depots "
            + " ".join(d for d in depots if d not in missing))

    units = []
    for depot in depots:
        units.extend(load_depot_lines(root / depot / f"{args.date}{suffix}", depot))
    print(f"Linek: {sum(len(u['line_ids']) for u in units)} "
          f"({len(units)} jednotek po sloučení dvojlinek) "
          f"za depa {', '.join(depots)}")

    familiarity = load_familiarity()
    result = build_assignment(units, registry, args.date,
                              familiarity=familiarity)

    rows = result_rows(result)
    print(format_table(rows))
    print()
    for w in result["warnings"]:
        print(w)

    out_dir = root / "driver_assignment" / f"{args.date}{suffix}"
    day_csv = out_dir / f"driver_plan_{args.date}.csv"
    write_csv(day_csv, rows)
    for depot in depots:
        depot_rows = [r for r in rows if r["depot"] == depot]
        if depot_rows:
            write_csv(root / depot / f"{args.date}{suffix}"
                      / f"driver_plan_{depot}_{args.date}.csv", depot_rows)
    (out_dir / "summary.json").write_text(json.dumps({
        "date": args.date, "label": args.label, "depots": depots,
        "registry_file": registry_path.name,
        "lines": sum(len(u["line_ids"]) for u in units),
        "units": len(units), "assigned": len(result["assigned"]),
        "uncovered": [{"depot": u["depot"], "lines": u["line_ids"],
                       "type": u["type_code"]} for u in result["uncovered"]],
        "weights": result["weights"], "warnings": result["warnings"],
        "params": {k: CONFIG[k] for k in
                   ("tight_slack_min", "tight_pos_coef", "quality_speed")},
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nVýstup: {day_csv.as_posix()} (+ driver_plan per depo)")
    if result["uncovered"]:
        sys.exit(2)


if __name__ == "__main__":
    main()
