"""
order_history.py — šance závozu z historie objednávek
======================================================

Predikční režim doplňuje do RiRo objednávky z minulých závozů („dopredikované").
Tenhle modul spočítá, s jakou pravděpodobností daná adresa v daný den v týdnu
opravdu objedná, a hodí za ni kostkou. Objednávka se pak do plánu dostane
CELÁ, nebo vůbec — kilogramy se nijak neškálují.

Šance = zavezené dny / způsobilé dny stejného dne v týdnu:

  okno    od prvního závozu po poslední pauze delší než 2 měsíce,
          nejvýš rok zpět; končí posledním dnem, který historie pokrývá
          (dny za koncem dat nejsou „nezavezeno", ale „nevíme")
  svátky  se vynechávají z čitatele i jmenovatele — jel-li někdo výjimečně
          na Velký pátek, nemá si tím zvednout ani snížit statistiku
  prázdno bez historie nebo bez způsobilého dne v okně = 100 % (radši
          naplánovat auto navíc než zákazníka vynechat)

Los je deterministický (seed = datum závozu + adresa), takže opakovaný běh
dá stejný plán a porovnání predikce s realitou má smysl.

Zdroj: data/historie_objednavky/*.xlsx — sloupce `Datum` a `Zkratka`
(= location_code v RiRo). Ostatní sloupce se ignorují, kg ani časy nás nezajímají.
"""

from __future__ import annotations

import random
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Iterable, Sequence

HISTORY_DIR = Path("data") / "historie_objednavky"
HISTORY_SHEET_FALLBACK = "List1"

# Jak daleko zpět se koukáme, i kdyby adresa jezdila roky.
WINDOW_DAYS = 365

# Mezera mezi dvěma závozy, po které bereme zákazníka jako „nového".
# Delší pauza znamená, že starší historie o dnešním chování nic neříká.
PAUSE_GAP_DAYS = 61          # ~2 měsíce; přesně 61 dní pauza ještě NENÍ

# Názvy sloupců v xlsx (hledají se v hlavičce, ne podle indexu).
COL_DATE_NAME     = "Datum"
COL_LOCATION_NAME = "Zkratka"

# České státní svátky s pevným datem (měsíc, den).
FIXED_HOLIDAYS = (
    (1, 1),    # Nový rok / Den obnovy samostatného českého státu
    (5, 1),    # Svátek práce
    (5, 8),    # Den vítězství
    (7, 5),    # Cyril a Metoděj
    (7, 6),    # Jan Hus
    (9, 28),   # Den české státnosti
    (10, 28),  # Vznik samostatného Československa
    (11, 17),  # Den boje za svobodu a demokracii
    (12, 24),  # Štědrý den
    (12, 25),  # 1. svátek vánoční
    (12, 26),  # 2. svátek vánoční
)


# ═════════════════════════════════════════════════════════════════════════════
#  Svátky
# ═════════════════════════════════════════════════════════════════════════════

def easter_sunday(year: int) -> date:
    """Velikonoční neděle podle anonymního gregoriánského algoritmu."""
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    lam = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * lam) // 451
    month, day = divmod(h + lam - 7 * m + 114, 31)
    return date(year, month, day + 1)


@lru_cache(maxsize=None)
def czech_holidays(year: int) -> frozenset[date]:
    """Všech 13 českých státních svátků daného roku (vč. Velikonoc)."""
    easter = easter_sunday(year)
    days = {date(year, month, day) for month, day in FIXED_HOLIDAYS}
    days.add(easter - timedelta(days=2))   # Velký pátek
    days.add(easter + timedelta(days=1))   # Velikonoční pondělí
    return frozenset(days)


def is_holiday(day: date) -> bool:
    return day in czech_holidays(day.year)


# ═════════════════════════════════════════════════════════════════════════════
#  Načtení historie
# ═════════════════════════════════════════════════════════════════════════════

@dataclass(frozen=True)
class History:
    """
    by_location  … location_code (lowercase) -> seřazené UNIKÁTNÍ dny závozu
    max_date     … poslední den, který data pokrývají (GLOBÁLNĚ, ne per adresa) —
                   za ním už nevíme, jestli se nejelo, nebo jen chybí data
    """
    by_location: dict[str, tuple[date, ...]]
    max_date: date | None
    files: tuple[str, ...]

    def dates_for(self, location_code: str) -> tuple[date, ...]:
        return self.by_location.get(str(location_code).strip().lower(), ())


def find_history_files(history_dir: Path = HISTORY_DIR) -> list[Path]:
    """Všechny xlsx s historií, seřazené podle názvu (2025.xlsx, 2026.xlsx …)."""
    if not history_dir.exists():
        raise FileNotFoundError(
            f"[CHYBA] Složka s historií objednávek neexistuje: {history_dir}\n"
            f"        Predikce z ní počítá šanci závozu. Vlož tam roční exporty "
            f"(např. 2025.xlsx, 2026.xlsx) se sloupci "
            f"'{COL_DATE_NAME}' a '{COL_LOCATION_NAME}'."
        )
    files = sorted(p for p in history_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not files:
        raise FileNotFoundError(
            f"[CHYBA] V {history_dir} není žádný .xlsx s historií objednávek.\n"
            f"        Predikce z ní počítá šanci závozu."
        )
    return files


def _parse_history_date(raw) -> date | None:
    """Datum z xlsx — openpyxl vrací datetime, nebo text 'YYYY-MM-DD HH:MM:SS.fff'."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def load_history(paths: Sequence[Path]) -> History:
    """
    Sloučí xlsx do mapy adresa -> dny závozu. Sloupce se hledají podle názvu
    v hlavičce, ať přežijeme přeházení pořadí.
    """
    import openpyxl                       # těžký import jen pro predikci

    by_location: dict[str, set[date]] = {}
    max_date: date | None = None

    for path in paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[HISTORY_SHEET_FALLBACK] if HISTORY_SHEET_FALLBACK in wb.sheetnames \
                else wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            names = [str(h).strip() if h is not None else "" for h in header]
            try:
                i_date = names.index(COL_DATE_NAME)
                i_loc  = names.index(COL_LOCATION_NAME)
            except ValueError:
                raise ValueError(
                    f"[CHYBA] {path.name} nemá v hlavičce sloupce "
                    f"'{COL_DATE_NAME}' a '{COL_LOCATION_NAME}'.\n"
                    f"        Nalezeno: {', '.join(n for n in names if n) or '(prázdná hlavička)'}"
                )
            for row in rows:
                if len(row) <= max(i_date, i_loc):
                    continue
                day = _parse_history_date(row[i_date])
                loc = row[i_loc]
                if day is None or loc is None:
                    continue
                key = str(loc).strip().lower()
                if not key:
                    continue
                by_location.setdefault(key, set()).add(day)
                if max_date is None or day > max_date:
                    max_date = day
        finally:
            wb.close()

    return History(
        by_location={k: tuple(sorted(v)) for k, v in by_location.items()},
        max_date=max_date,
        files=tuple(p.name for p in paths),
    )


# ═════════════════════════════════════════════════════════════════════════════
#  Výpočet šance
# ═════════════════════════════════════════════════════════════════════════════

def _window_start(dates: Sequence[date], target_date: date, window_end: date) -> tuple[date, bool]:
    """
    Začátek okna = první závoz po POSLEDNÍ pauze delší než PAUSE_GAP_DAYS,
    nejvýš však rok zpět. Vrací (začátek, jestli se pauza uplatnila).
    """
    earliest = target_date - timedelta(days=WINDOW_DAYS)
    start = dates[0]
    pause_applied = False
    for prev, curr in zip(dates, dates[1:]):
        if (curr - prev).days > PAUSE_GAP_DAYS:
            start = curr
            pause_applied = True
    return max(start, earliest), pause_applied and start > earliest


def _count_eligible_days(start: date, end: date, weekday: int) -> int:
    """Kolik dní daného dne v týdnu je v intervalu, po odečtení svátků."""
    if start > end:
        return 0
    first = start + timedelta(days=(weekday - start.weekday()) % 7)
    count = 0
    day = first
    while day <= end:
        if not is_holiday(day):
            count += 1
        day += timedelta(days=7)
    return count


def delivery_chance(dates: Sequence[date], target_date: date, *,
                    history_max_date: date | None) -> dict:
    """
    Šance, že adresa objedná v `target_date`, spočítaná ze stejných dnů v týdnu.

    Bez historie nebo bez jediného způsobilého dne v okně vrací 1.0 — plán
    radši pojede s objednávkou navíc, než abychom zákazníka tiše vynechali.
    """
    weekday = target_date.weekday()
    window_end = target_date - timedelta(days=1)
    if history_max_date is not None and history_max_date < window_end:
        window_end = history_max_date

    usable = [d for d in sorted(set(dates)) if d <= window_end]
    if not usable:
        return {
            "chance": 1.0, "delivered": 0, "eligible": 0,
            "window_start": None, "window_end": window_end,
            "weekday": weekday, "no_history": True, "pause_applied": False,
        }

    window_start, pause_applied = _window_start(usable, target_date, window_end)
    delivered = sum(
        1 for d in usable
        if window_start <= d <= window_end and d.weekday() == weekday and not is_holiday(d)
    )
    eligible = _count_eligible_days(window_start, window_end, weekday)
    chance = 1.0 if eligible == 0 else delivered / eligible

    return {
        "chance": chance, "delivered": delivered, "eligible": eligible,
        "window_start": window_start, "window_end": window_end,
        "weekday": weekday, "no_history": False, "pause_applied": pause_applied,
    }


# ═════════════════════════════════════════════════════════════════════════════
#  Los a vyhodnocení
# ═════════════════════════════════════════════════════════════════════════════

def prediction_draw(date_str: str, location_code: str) -> float:
    """
    Deterministický los z (datum závozu, adresa) — opakovaný běh dá stejný plán.

    Všechny objednávky téže adresy v týž den sdílejí jeden los. Je to záměr:
    adresa buď ten den objedná, nebo ne, a nemá smysl losovat každou položku zvlášť.
    """
    seed = f"{date_str}|{str(location_code).strip().lower()}"
    return random.Random(seed).random()


def evaluate_predictions(predicted_rows: Iterable[dict], history: History,
                         date_str: str) -> list[dict]:
    """Pro každý dopredikovaný RiRo řádek spočítá šanci a hodí kostkou."""
    target_date = date.fromisoformat(date_str)
    records = []
    for raw in predicted_rows:
        location = str(raw.get("location_code", "")).strip()
        info = delivery_chance(history.dates_for(location), target_date,
                               history_max_date=history.max_date)
        draw = prediction_draw(date_str, location)
        records.append({
            "line": raw.get("_line"),
            "order_number": raw.get("order_number", ""),
            "location_code": location,
            "customer_name": raw.get("customer_name", ""),
            "draw": draw,
            "selected": info["chance"] >= 1.0 or draw < info["chance"],
            **info,
        })
    return records


def format_prediction_report(records: Sequence[dict], date_str: str) -> str:
    """Tabulka do konzole: která dopredikovaná objednávka prošla losem a proč."""
    if not records:
        return "\nPredikce: žádné dopredikované objednávky.\n"

    weekday_names = ("pondělí", "úterý", "středa", "čtvrtek",
                     "pátek", "sobota", "neděle")
    target = date.fromisoformat(date_str)
    lines = [
        "",
        "=" * 78,
        f"PREDIKCE — šance závozu podle historie ({weekday_names[target.weekday()]} "
        f"{target.strftime('%d.%m.%Y')})",
        "=" * 78,
        f"{'adresa':<30} {'dny':>9} {'šance':>7} {'los':>7}  výsledek",
        "-" * 78,
    ]
    for r in sorted(records, key=lambda x: (not x["selected"], x["location_code"])):
        if r["no_history"]:
            ratio = "bez hist."
        elif r["eligible"] == 0:
            ratio = "0 dnů"
        else:
            ratio = f"{r['delivered']}/{r['eligible']}"
        pause = " *" if r["pause_applied"] else ""
        lines.append(
            f"{r['location_code'][:30]:<30} {ratio:>9} "
            f"{r['chance']:>6.0%} {r['draw']:>7.3f}  "
            f"{'VYBRÁNA' if r['selected'] else 'VYNECHÁNA'}{pause}"
        )
    included = sum(1 for r in records if r["selected"])
    lines += [
        "-" * 78,
        f"Dopredikovaných: {len(records)}  |  vybráno: {included}  |  "
        f"vynecháno: {len(records) - included}",
    ]
    if any(r["pause_applied"] for r in records):
        lines.append("* historie zkrácena kvůli pauze delší než dva měsíce")
    lines.append("=" * 78)
    return "\n".join(lines)


def build_prediction_stats(records: Sequence[dict], history: History) -> dict:
    """Blok `prediction` do prepare_stats JSON — ať je los dohledatelný zpětně."""
    included = sum(1 for r in records if r["selected"])
    return {
        "mode": "chance_from_history",
        "history_files": list(history.files),
        "history_max_date": history.max_date.isoformat() if history.max_date else None,
        "predicted_orders": len(records),
        "included": included,
        "skipped_by_chance": len(records) - included,
        "orders": [
            {
                "line": r["line"],
                "order_number": r["order_number"],
                "location_code": r["location_code"],
                "delivered": r["delivered"],
                "eligible": r["eligible"],
                "chance": round(r["chance"], 4),
                "draw": round(r["draw"], 4),
                "selected": r["selected"],
                "no_history": r["no_history"],
                "pause_applied": r["pause_applied"],
                "window_start": r["window_start"].isoformat() if r["window_start"] else None,
                "window_end": r["window_end"].isoformat() if r["window_end"] else None,
            }
            for r in records
        ],
    }
